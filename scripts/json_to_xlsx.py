#!/usr/bin/env python3
"""Convert Democracy Pulse JSON data to XLSX spreadsheet (backup database)."""

import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = Path(__file__).resolve().parent.parent

structure = {
    "Pillar I — Integrated Will": {
        "pillar_key": "pillar_I_integrated_will",
        "dimensions": {
            "government_type": "Government Type",
            "information_environment": "Information Environment",
            "economic_precarity": "Economic Precarity",
            "forced_substitution_coefficient": "Forced Substitution Coefficient",
            "civic_education": "Civic Education",
            "digital_rights": "Digital Rights",
        }
    },
    "Pillar II — Organic Ties": {
        "pillar_key": "pillar_II_organic_ties",
        "dimensions": {
            "social_cohesion": "Social Cohesion",
            "family_structure_health": "Family Structure Health",
            "civic_participation_beyond_voting": "Civic Participation Beyond Voting",
            "internal_dissent_capacity": "Internal Dissent Capacity",
            "social_isolation": "Social Isolation",
        }
    },
    "Pillar III — Social Capital Transformation": {
        "pillar_key": "pillar_III_social_capital_transformation",
        "dimensions": {
            "wealth_distribution": "Wealth Distribution",
            "taxation_progressivity": "Taxation Progressivity",
            "public_investment": "Public Investment",
            "social_mobility": "Social Mobility",
            "cultural_production_accessibility": "Cultural Production Accessibility",
            "corporate_extraction_vs_public_benefit": "Corporate Extraction vs Public Benefit",
        }
    }
}

INFO_ENV_SUBS = ["mainstream_media", "digital_media", "economic_structure", "social_media_journalism"]

COUNTRIES = [
    ("turkey", "Turkey"),
    ("usa", "United States"),
    ("norway", "Norway"),
]

HEADER_FONT = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")

PILLAR_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
PILLAR_FILL = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")

DIM_FONT = Font(name="Calibri", bold=True, size=10)
DIM_FILL = PatternFill(start_color="e0e0e0", end_color="e0e0e0", fill_type="solid")

SUB_FONT = Font(name="Calibri", bold=True, italic=True, size=10)
SUB_FILL = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")

COL_HEADER_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
COL_HEADER_FILL = PatternFill(start_color="0f3460", end_color="0f3460", fill_type="solid")

FIELD_FONT = Font(name="Calibri", size=10)
NULL_FONT = Font(name="Calibri", size=10, color="aaaaaa")

THIN_BORDER = Border(
    left=Side(style="thin", color="cccccc"),
    right=Side(style="thin", color="cccccc"),
    top=Side(style="thin", color="cccccc"),
    bottom=Side(style="thin", color="cccccc"),
)


def humanize(key: str) -> str:
    """Convert snake_case to Title Case."""
    return key.replace("_", " ").title()


def fmt_val(val):
    """Format a value for display."""
    if val is None:
        return "—"
    if isinstance(val, bool):
        return "Yes" if val else "No"
    if isinstance(val, (int, float)):
        return val
    return str(val)


def write_country(ws, data, country_name):
    """Write one country's data to a worksheet."""
    # Column widths
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20

    # Title
    row = 1
    ws.merge_cells("A1:C1")
    cell = ws["A1"]
    cell.value = f"Democracy Pulse — {country_name}"
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # Metadata
    row = 2
    ws.merge_cells("A2:C2")
    cell = ws["A2"]
    cell.value = f"Data collected: {data.get('collection_date', '—')} | Version: {data.get('data_version', '—')}"
    cell.font = Font(name="Calibri", size=9, color="999999")
    cell.alignment = Alignment(horizontal="center")

    # Column headers
    row = 3
    for col, label in enumerate(["Field", "Value", "Source"], 1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = COL_HEADER_FONT
        cell.fill = COL_HEADER_FILL
        cell.border = THIN_BORDER

    row = 4

    for pillar_name, pillar_cfg in structure.items():
        pkey = pillar_cfg["pillar_key"]
        pdata = data.get(pkey, {})

        # Pillar header
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws.cell(row=row, column=1, value=pillar_name)
        cell.font = PILLAR_FONT
        cell.fill = PILLAR_FILL
        cell.alignment = Alignment(horizontal="left")
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).fill = PILLAR_FILL
        row += 1

        for dim_key, dim_label in pillar_cfg["dimensions"].items():
            # Dimension header
            ws.merge_cells(f"A{row}:C{row}")
            cell = ws.cell(row=row, column=1, value=f"  {dim_label}")
            cell.font = DIM_FONT
            cell.fill = DIM_FILL
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = THIN_BORDER
                ws.cell(row=row, column=c).fill = DIM_FILL
            row += 1

            dim_data = pdata.get(dim_key, {})
            if not dim_data:
                continue

            if dim_key == "information_environment":
                # Has sub-sections
                for sub_key in INFO_ENV_SUBS:
                    sub_data = dim_data.get(sub_key, {})
                    if not sub_data:
                        continue
                    # Sub-section header
                    ws.merge_cells(f"A{row}:C{row}")
                    cell = ws.cell(row=row, column=1, value=f"    {humanize(sub_key)}")
                    cell.font = SUB_FONT
                    cell.fill = SUB_FILL
                    for c in range(1, 4):
                        ws.cell(row=row, column=c).border = THIN_BORDER
                        ws.cell(row=row, column=c).fill = SUB_FILL
                    row += 1

                    for fkey, fval in sub_data.items():
                        if fkey == "source":
                            continue
                        cell_a = ws.cell(row=row, column=1, value=f"      {humanize(fkey)}")
                        cell_a.font = FIELD_FONT
                        cell_a.border = THIN_BORDER
                        cell_b = ws.cell(row=row, column=2, value=fmt_val(fval))
                        cell_b.font = FIELD_FONT if fval is not None else NULL_FONT
                        cell_b.border = THIN_BORDER
                        cell_c = ws.cell(row=row, column=3, value=sub_data.get("source", ""))
                        cell_c.font = FIELD_FONT
                        cell_c.border = THIN_BORDER
                        row += 1
            else:
                # Flat dimension
                for fkey, fval in dim_data.items():
                    if fkey == "source":
                        continue
                    if isinstance(fval, dict):
                        continue  # skip nested objects
                    cell_a = ws.cell(row=row, column=1, value=f"    {humanize(fkey)}")
                    cell_a.font = FIELD_FONT
                    cell_a.border = THIN_BORDER
                    cell_b = ws.cell(row=row, column=2, value=fmt_val(fval))
                    cell_b.font = FIELD_FONT if fval is not None else NULL_FONT
                    cell_b.border = THIN_BORDER
                    cell_c = ws.cell(row=row, column=3, value=dim_data.get("source", ""))
                    cell_c.font = FIELD_FONT
                    cell_c.border = THIN_BORDER
                    row += 1


def main():
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for cidx, (ckey, cname) in enumerate(COUNTRIES):
        fpath = BASE / "data" / "raw" / f"{ckey}_pilot_v0.1.json"
        with open(fpath) as f:
            data = json.load(f)

        ws = wb.create_sheet(title=cname)
        write_country(ws, data, cname)

    # Also create a "Comparison" sheet with all countries side by side
    ws_comp = wb.create_sheet(title="Comparison")
    ws_comp.column_dimensions["A"].width = 38
    ws_comp.column_dimensions["B"].width = 22
    ws_comp.column_dimensions["C"].width = 22
    ws_comp.column_dimensions["D"].width = 22

    # Title
    ws_comp.merge_cells("A1:D1")
    cell = ws_comp["A1"]
    cell.value = "Democracy Pulse — Country Comparison"
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    ws_comp.row_dimensions[1].height = 30

    # Column headers
    row = 2
    for col, label in enumerate(["Field", "Turkey", "United States", "Norway"], 1):
        cell = ws_comp.cell(row=row, column=col, value=label)
        cell.font = COL_HEADER_FONT
        cell.fill = COL_HEADER_FILL
        cell.border = THIN_BORDER
    row = 3

    all_data = {}
    for ckey, cname in COUNTRIES:
        fpath = BASE / "data" / "raw" / f"{ckey}_pilot_v0.1.json"
        with open(fpath) as f:
            all_data[ckey] = json.load(f)

    for pillar_name, pillar_cfg in structure.items():
        pkey = pillar_cfg["pillar_key"]

        # Pillar header
        ws_comp.merge_cells(f"A{row}:D{row}")
        cell = ws_comp.cell(row=row, column=1, value=pillar_name)
        cell.font = PILLAR_FONT
        cell.fill = PILLAR_FILL
        for c in range(1, 5):
            ws_comp.cell(row=row, column=c).border = THIN_BORDER
            ws_comp.cell(row=row, column=c).fill = PILLAR_FILL
        row += 1

        for dim_key, dim_label in pillar_cfg["dimensions"].items():
            # Dimension header
            ws_comp.merge_cells(f"A{row}:D{row}")
            cell = ws_comp.cell(row=row, column=1, value=f"  {dim_label}")
            cell.font = DIM_FONT
            cell.fill = DIM_FILL
            for c in range(1, 5):
                ws_comp.cell(row=row, column=c).border = THIN_BORDER
                ws_comp.cell(row=row, column=c).fill = DIM_FILL
            row += 1

            for ckey in ["turkey", "usa", "norway"]:
                dim_data = all_data[ckey].get(pkey, {}).get(dim_key, {})
                if dim_key == "information_environment":
                    for sub_key in INFO_ENV_SUBS:
                        sub_data = dim_data.get(sub_key, {})
                        if not sub_data:
                            continue
                        # Sub header
                        ws_comp.merge_cells(f"A{row}:D{row}")
                        cell = ws_comp.cell(row=row, column=1, value=f"    {humanize(sub_key)}")
                        cell.font = SUB_FONT
                        cell.fill = SUB_FILL
                        for c in range(1, 5):
                            ws_comp.cell(row=row, column=c).border = THIN_BORDER
                            ws_comp.cell(row=row, column=c).fill = SUB_FILL
                        row += 1

                        for fkey in sub_data:
                            if fkey == "source":
                                continue
                            cell_a = ws_comp.cell(row=row, column=1, value=f"      {humanize(fkey)}")
                            cell_a.font = FIELD_FONT
                            cell_a.border = THIN_BORDER
                            # Get value for each country
                            for ci, ck in enumerate(["turkey", "usa", "norway"]):
                                sv = all_data[ck].get(pkey, {}).get(dim_key, {}).get(sub_key, {}).get(fkey)
                                cell_v = ws_comp.cell(row=row, column=2 + ci, value=fmt_val(sv))
                                cell_v.font = FIELD_FONT if sv is not None else NULL_FONT
                                cell_v.border = THIN_BORDER
                            row += 1
                    break  # Only iterate sub-sections once
                else:
                    # Flat fields
                    for fkey, fval in dim_data.items():
                        if fkey == "source" or isinstance(fval, dict):
                            continue
                        cell_a = ws_comp.cell(row=row, column=1, value=f"    {humanize(fkey)}")
                        cell_a.font = FIELD_FONT
                        cell_a.border = THIN_BORDER
                        for ci, ck in enumerate(["turkey", "usa", "norway"]):
                            fv = all_data[ck].get(pkey, {}).get(dim_key, {}).get(fkey)
                            cell_v = ws_comp.cell(row=row, column=2 + ci, value=fmt_val(fv))
                            cell_v.font = FIELD_FONT if fv is not None else NULL_FONT
                            cell_v.border = THIN_BORDER
                        row += 1
                    break  # Only list fields once

    out = BASE / "data" / "democracy_pulse_backup.xlsx"
    wb.save(str(out))
    print(f"Saved: {out}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Rows in Turkey sheet: ~{wb['Turkey'].max_row}")


if __name__ == "__main__":
    main()